
import streamlit as st
import pandas as pd
import numpy as np
import datetime as dt
from io import BytesIO
from openpyxl import load_workbook
import math
import re

st.set_page_config(page_title="Cosmic Generator", layout="wide")

# ---------- Swiss Ephemeris (optional) ----------
try:
    import swisseph as swe
    HAVE_SWE = True
except Exception:
    HAVE_SWE = False

# ---------- Helpers ----------
@st.cache_data(show_spinner=False)
def load_workbook_bytes(b: bytes):
    return load_workbook(filename=BytesIO(b), data_only=True)

@st.cache_data(show_spinner=False)
def load_default_workbook():
    with open("data/cosmic_generator_v25.xlsx", "rb") as f:
        return f.read()

def get_sheet_df(wb, name):
    try:
        if not wb or name not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb[name]
        rows = list(ws.values)
        if not rows:
            return pd.DataFrame()
        header = rows[0]
        df = pd.DataFrame(rows[1:], columns=header)
        return df.dropna(how="all")
    except Exception:
        return pd.DataFrame()

ZODIAC = ["Aries","Taurus","Gemini","Cancer","Leo","Virgo","Libra","Scorpio","Sagittarius","Capricorn","Aquarius","Pisces"]

def sun_sign_from_date(d: dt.date):
    m, day = d.month, d.day
    if   (m==3 and day>=21) or (m==4 and day<=19): return "Aries"
    elif (m==4 and day>=20) or (m==5 and day<=20): return "Taurus"
    elif (m==5 and day>=21) or (m==6 and day<=20): return "Gemini"
    elif (m==6 and day>=21) or (m==7 and day<=22): return "Cancer"
    elif (m==7 and day>=23) or (m==8 and day<=22): return "Leo"
    elif (m==8 and day>=23) or (m==9 and day<=22): return "Virgo"
    elif (m==9 and day>=23) or (m==10 and day<=22): return "Libra"
    elif (m==10 and day>=23) or (m==11 and day<=21): return "Scorpio"
    elif (m==11 and day>=22) or (m==12 and day<=21): return "Sagittarius"
    elif (m==12 and day>=22) or (m==1 and day<=19): return "Capricorn"
    elif (m==1 and day>=20) or (m==2 and day<=18): return "Aquarius"
    elif (m==2 and day>=19) or (m==3 and day<=20): return "Pisces"
    return ""

def _deg_to_rad(x): return x * math.pi / 180.0
def _rev(x): return x % 360.0

def moon_longitude_approx_noon_utc(d: dt.date):
    year, month, day = d.year, d.month, d.day
    hour = 12.0
    if month <= 2:
        year -= 1
        month += 12
    A = int(year/100)
    B = 2 - A + int(A/4)
    JD = int(365.25*(year + 4716)) + int(30.6001*(month + 1)) + (day + hour/24.0) + B - 1524.5
    T = (JD - 2451545.0)/36525.0
    Lp = _rev(218.3164477 + 481267.88123421*T - 0.0015786*T*T)
    D  = _rev(297.8501921 + 445267.1114034*T - 0.0018819*T*T)
    M  = _rev(357.5291092 + 35999.0502909*T - 0.0001536*T*T)
    Mp = _rev(134.9633964 + 477198.8675055*T + 0.0087414*T*T)
    F  = _rev(93.2720950 + 483202.0175233*T - 0.0036539*T*T)
    Dr = _deg_to_rad(D); Mr = _deg_to_rad(M); Mpr = _deg_to_rad(Mp); Fr = _deg_to_rad(F)
    lon = (Lp
        + 6.289 * math.sin(Mpr)
        + 1.274 * math.sin(2*Dr - Mpr)
        + 0.658 * math.sin(2*Dr)
        + 0.214 * math.sin(2*Mpr)
        - 0.186 * math.sin(Mr)
        - 0.114 * math.sin(2*Fr)
        + 0.059 * math.sin(2*Dr - 2*Mpr)
        + 0.057 * math.sin(2*Dr - Mr - Mpr)
        + 0.053 * math.sin(2*Dr + Mpr)
        + 0.046 * math.sin(2*Dr - Mr)
        + 0.041 * math.sin(Mr + Mpr))
    return _rev(lon)

def moon_sign_exact(birthdate: dt.date, birthtime: dt.time, tz_offset: float):
    try:
        hour_local = birthtime.hour + birthtime.minute/60 + birthtime.second/3600
        hour_utc = hour_local - float(tz_offset)
        jd = swe.julday(birthdate.year, birthdate.month, birthdate.day, hour_utc) if HAVE_SWE else None
        if HAVE_SWE and jd is not None:
            pos = swe.calc_ut(jd, swe.MOON)[0]
            lon = float(pos[0])
            idx = int(lon // 30) % 12
            return ZODIAC[idx], lon, True
        else:
            lon = moon_longitude_approx_noon_utc(birthdate)
            idx = int(lon // 30) % 12
            return ZODIAC[idx], lon, False
    except Exception:
        lon = moon_longitude_approx_noon_utc(birthdate)
        idx = int(lon // 30) % 12
        return ZODIAC[idx], lon, False

def universal_day_number(d: dt.date, keep_master=True):
    s = d.strftime("%Y%m%d")
    n = sum(int(ch) for ch in s)
    if keep_master and n in (11, 22, 33):
        return n
    return 9 if n % 9 == 0 else n % 9

def weekday_name(d: dt.date):
    return d.strftime("%A")

def safe_unique_list(series):
    try:
        return sorted([x for x in series.dropna().unique().tolist() if x])
    except Exception:
        return []

# ---------- Sidebar: data source ----------
st.sidebar.title("Data Source")
uploaded = st.sidebar.file_uploader("Upload a Cosmic Generator workbook (.xlsx)", type=["xlsx"], help="Upload latest vXX to use it live.")
keep_master = st.sidebar.checkbox("Numerology: Keep master numbers (11/22/33)?", value=True)
st.session_state["keep_master"] = keep_master

raw = uploaded.getvalue() if uploaded else load_default_workbook()
wb = load_workbook_bytes(raw) if raw else None

# ---------- Load sheets ----------
df_data         = get_sheet_df(wb, "Data")
df_audit        = get_sheet_df(wb, "AuditData")
df_elem_items   = get_sheet_df(wb, "Element_Items")
df_zones        = get_sheet_df(wb, "House_Zones")
df_rel          = get_sheet_df(wb, "Element_Relations")
df_pref         = get_sheet_df(wb, "Element_Preferences")
df_shapes       = get_sheet_df(wb, "Shape_Elements")
df_guide        = get_sheet_df(wb, "Activity_Day_Guide")

# Master lookup helper
def get_sign_row(sign):
    if df_data.empty or "Astrological Sign" not in df_data.columns:
        return pd.Series(dtype=object)
    row = df_data.loc[df_data["Astrological Sign"]==sign]
    return row.iloc[0] if not row.empty else pd.Series(dtype=object)

# ---------- Tabs ----------
tabs = st.tabs(["Inputs", "Life Audit", "Activity Timing", "House Zone Checker", "Items Browser"])

# ===== Inputs =====
with tabs[0]:
    st.subheader("Inputs")
    birthdate = st.date_input("Birthdate", dt.date(1990,1,1))
    birthtime = st.time_input("Birth time (local)", dt.time(12,0))

    tz_options = [
        ("UTC (0)", 0.0),
        ("London Winter (0)", 0.0),
        ("London Summer (+1)", 1.0),
        ("Paris / Berlin (+1 winter / +2 summer)", 1.0),
        ("Athens (+2)", 2.0),
        ("Moscow (+3)", 3.0),
        ("Dubai (+4)", 4.0),
        ("India IST (+5.5)", 5.5),
        ("Bangladesh (+6)", 6.0),
        ("Thailand (+7)", 7.0),
        ("Singapore / Hong Kong (+8)", 8.0),
        ("Japan (+9)", 9.0),
        ("Sydney (+10)", 10.0),
        ("Auckland (+12)", 12.0),
        ("New York (-5 winter / -4 summer)", -5.0),
        ("Chicago (-6 winter / -5 summer)", -6.0),
        ("Denver (-7 winter / -6 summer)", -7.0),
        ("Los Angeles (-8 winter / -7 summer)", -8.0),
        ("Mexico City (-6)", -6.0),
        ("Sao Paulo (-3)", -3.0),
        ("Johannesburg (+2)", 2.0),
        ("Cairo (+2)", 2.0),
        ("Tehran (+3.5)", 3.5),
        ("Kathmandu (+5.75)", 5.75),
        ("Yangon (+6.5)", 6.5),
    ]
    tz_labels = [x[0] for x in tz_options]
    tz_map = {k:v for k,v in tz_options}
    tz_choice = st.selectbox("Timezone (pick closest)", tz_labels, index=7)
    tz_offset = tz_map.get(tz_choice, 0.0)

    sun_sign = sun_sign_from_date(birthdate) or ""
    moon_sign, moon_lon, used_exact = moon_sign_exact(birthdate, birthtime, tz_offset)
    st.info(f"Sun: {sun_sign or '—'} | Moon: {moon_sign or '—'} {'(exact)' if used_exact else '(approx)'} | TZ: {tz_choice}")

    use_choice = st.radio("Use which sign across the app?", ["Sun", "Moon"], horizontal=True)
    selected_sign = sun_sign if use_choice == "Sun" else moon_sign
    st.session_state["selected_sign"] = selected_sign

    # Full details for selected sign (from Data sheet)
    if not df_data.empty and "Astrological Sign" in df_data.columns and selected_sign:
        row = df_data.loc[df_data["Astrological Sign"] == selected_sign]
        if not row.empty:
            info = row.iloc[0].to_dict()
            info_clean = {k: v for k,v in info.items() if pd.notna(v) and str(v) != ""}
            df_show = pd.DataFrame(list(info_clean.items()), columns=["Field","Value"])
            st.write("**Sign Details (from Data sheet):**")
            st.dataframe(df_show, use_container_width=True, hide_index=True)

# ===== Life Audit (v8 Fast Track) =====
with tabs[1]:
    st.subheader("Life Audit – Conflicts & Remedies (v8)")
    selected_sign = st.session_state.get("selected_sign","")
    if not selected_sign:
        st.warning("Go to Inputs and choose Sun or Moon first.")
    st.caption("Add multiple items per box (comma-separated). We'll flag STRONG/MILD conflicts and suggest top fixes.")

    # Input categories (textarea for multiple tokens)
    config = [
        ("Colours / Décor", "Avoid Household (strong)", "Avoid Household (mild)", "Colour", "Household Items"),
        ("Foods", "Avoid Foods (strong)", "Avoid Foods (mild)", "Foods", "Foods"),
        ("Crystals & Gemstones", "Avoid Crystals (strong)", "Avoid Crystals (mild)", "Primary Crystals", "Alternative Crystals / Gemstones"),
        ("Activities", "Avoid Activities (strong)", "Avoid Activities (mild)", "Favorable Activities", "Favorable Activities"),
        ("Elements", "Avoid Elements (strong)", "Avoid Elements (mild)", "Element", "Element"),
        ("People (Signs)", "Enemy Signs (strong)", "Enemy Signs (mild)", None, None),
    ]

    # Simple synonym map & normalizer
    SYN = {
        "sofa":"sofa","couch":"sofa",
        "fridge":"fridge","refrigerator":"fridge","freezer":"fridge",
        "phone":"phone","mobile":"phone","cell phone":"phone","cellphone":"phone",
        "grey":"gray","rose gold":"gold",
        "laptop":"computer","pc":"computer","desktop":"computer",
        "trash":"bin","trash can":"bin","garbage":"bin","garbage can":"bin","dustbin":"bin",
        "loo":"toilet","wc":"toilet","lavatory":"toilet",
        "aqua":"water",
    }
    def normalize_token(t):
        t = re.sub(r"[^a-z0-9\s\-]", "", t.lower()).strip()
        t = re.sub(r"\s+", " ", t)
        t = SYN.get(t, t)
        # simple singular
        if t.endswith("es"): t = t[:-2]
        elif t.endswith("s") and len(t)>3: t = t[:-1]
        return t

    def split_tokens(s):
        if not s: return []
        # split by comma or newline
        raw = re.split(r"[,\\n]", s)
        tokens = [normalize_token(x) for x in raw if x and x.strip()]
        # also split on spaces for multi-word variants (keep originals too)
        extras = []
        for t in tokens:
            parts = t.split(" ")
            if len(parts)>1:
                extras.extend([p for p in parts if p])
        return list(dict.fromkeys(tokens + extras))  # unique, keep order

    # Build audit lookup for chosen sign
    def lookup_avoid(col):
        if df_audit.empty or "Astrological Sign" not in df_audit.columns or col not in df_audit.columns:
            return []
        row = df_audit.loc[df_audit["Astrological Sign"]==selected_sign]
        if row.empty: return []
        raw = str(row.iloc[0][col] or "")
        toks = [normalize_token(x) for x in re.split(r",", raw) if x and x.strip()]
        return list(dict.fromkeys(toks))

    results_rows = []
    summary_rows = []

    # Create 2-column layout per row of categories
    for idx, (label,strong_col,mild_col,rem1,rem2) in enumerate(config):
        st.markdown(f"**{label}**")
        user_text = st.text_area(f"My {label} (comma-separated)", key=f"la8_{idx}", height=60, placeholder="e.g., blue, rose gold, marble")
        tokens = split_tokens(user_text)

        strong_set = set(lookup_avoid(strong_col))
        mild_set   = set(lookup_avoid(mild_col))

        strong_hits = 0
        mild_hits = 0

        for tok in tokens or [""]:
            if not tok:
                continue
            verdict, matched = "OK", ""
            # exact
            if tok in strong_set:
                verdict, matched = "STRONG", tok
            elif tok in mild_set:
                verdict, matched = "MILD", tok
            else:
                # substring match
                if any(tok and tok in x for x in strong_set):
                    verdict, matched = "STRONG", next(x for x in strong_set if tok in x)
                elif any(tok and tok in x for x in mild_set):
                    verdict, matched = "MILD", next(x for x in mild_set if tok in x)
            if verdict=="STRONG": strong_hits += 1
            elif verdict=="MILD": mild_hits += 1
            results_rows.append((label, tok, verdict, matched))

        # Priority remedies (Top 3) for this category
        fixes = []
        sign_row = get_sign_row(selected_sign)
        if strong_hits>0:
            if label=="Colours / Décor":
                if "Colour" in sign_row:
                    fixes.append(f"Switch to favourable colours: {sign_row['Colour']}")
                if "Shape" in sign_row:
                    fixes.append(f"Favor shapes: {sign_row['Shape']}")
                if "Element" in sign_row:
                    fixes.append(f"Use décor that expresses {sign_row['Element']} element")
            elif label=="Foods":
                if "Foods" in sign_row:
                    fixes.append(f"Prioritize: {sign_row['Foods']}")
                if "Avoid Foods (mild)" in df_audit.columns:
                    fixes.append("Reduce flagged foods; phase out strong conflicts first")
                fixes.append("Hydration and simple, whole foods for 7 days")
            elif label=="Crystals & Gemstones":
                if "Primary Crystals" in sign_row:
                    fixes.append(f"Carry/wear: {sign_row['Primary Crystals']}")
                if "Alternative Crystals / Gemstones" in sign_row:
                    fixes.append(f"Alternate set: {sign_row['Alternative Crystals / Gemstones']}")
                fixes.append("Cleanse & charge crystals weekly")
            elif label=="Activities":
                if "Favorable Activities" in sign_row:
                    fixes.append(f"Swap to: {sign_row['Favorable Activities']}")
                fixes.append("Schedule during your good weekday or number (see Activity Timing)")
            elif label=="Elements":
                if "Element" in sign_row:
                    fixes.append(f"Emphasize {sign_row['Element']} items; avoid opposing element clusters")
                fixes.append("Add 1-2 supportive items (colour/crystal) before removing big items")
            elif label=="People (Signs)":
                fixes.append("Reduce exposure to strong enemy-sign dynamics")
                fixes.append("Set clear boundaries; favor cooperative signs from your Data sheet if listed")
                fixes.append("Choose neutral/shared‑element activities")

        elif mild_hits>0:
            if label in ("Colours / Décor","Elements","Crystals & Gemstones"):
                fixes.append("Balance with supportive colours/crystals of your element")
            if label=="Foods":
                fixes.append("Moderate intake; avoid combos with other flagged foods")
            if label=="Activities":
                fixes.append("Do on a favourable weekday/number to offset mild conflicts")
            if label=="People (Signs)":
                fixes.append("Choose neutral settings; short interactions")

        # keep only top 3
        fixes = [f for f in fixes if f][:3]
        summary_rows.append((label, strong_hits, mild_hits, ", ".join(fixes) if fixes else ""))

    # Show detailed token results and category summary
    colA, colB = st.columns([2,1])
    with colA:
        df_res = pd.DataFrame(results_rows, columns=["Category","Token","Conflict","Matched Term"])
        st.write("**Token-by-token results**")
        st.dataframe(df_res, use_container_width=True)
    with colB:
        df_sum = pd.DataFrame(summary_rows, columns=["Category","STRONG hits","MILD hits","Top fixes"])
        st.write("**Category summary**")
        st.dataframe(df_sum, use_container_width=True)

# ===== Activity Timing =====
with tabs[2]:
    st.subheader("Activity Timing – Astrology x Numerology")
    if df_guide.empty:
        st.warning("Activity_Day_Guide sheet not found in workbook.")
    else:
        activities = safe_unique_list(df_guide["Activity"]) if "Activity" in df_guide.columns else []
        activity = st.selectbox("Activity", activities) if activities else ""
        date = st.date_input("Planned Date", dt.date.today())
        if activity:
            row = df_guide.loc[df_guide["Activity"]==activity]
            if row.empty:
                st.warning("Selected activity not found in guide."); 
            else:
                row = row.iloc[0]
                wd = weekday_name(date)
                ud = universal_day_number(date, keep_master=st.session_state.get("keep_master", True))
                good_days = str(row.get("Good Days (Astrology)",""))
                avoid_days = str(row.get("Avoid Days (Astrology)",""))
                good_nums = str(row.get("Good Numbers (Numerology)",""))
                avoid_nums = str(row.get("Avoid Numbers (Numerology)",""))
                notes = str(row.get("Synergy Notes",""))
                afit = "Yes" if wd in good_days else ("No" if wd in avoid_days else "Maybe")
                nfit = "Yes" if str(ud) in good_nums else ("No" if str(ud) in avoid_nums else "Maybe")
                verdict = "Strong Cosmic Timing" if (afit=="Yes" and nfit=="Yes") else ("Weak / Reschedule Suggested" if ("No" in (afit,nfit)) else "Moderate Timing")
                st.write(f"Weekday: {wd}   |   Universal Day: {ud}")
                st.write(f"Astrology Fit: {afit}   |   Numerology Fit: {nfit}")
                st.write(f"Overall Verdict: {verdict}")
                if notes:
                    st.info(notes)

# ===== House Zone Checker =====
with tabs[3]:
    st.subheader("House Zone Checker – 5 Elements + Space (with Shapes)")
    if df_elem_items.empty or df_zones.empty or df_rel.empty:
        st.warning("Missing one of: Element_Items, House_Zones, Element_Relations sheets.")
    else:
        items = safe_unique_list(df_elem_items["Item Name"]) if "Item Name" in df_elem_items.columns else []
        zones = safe_unique_list(df_zones["Zone"]) if "Zone" in df_zones.columns else []
        shapes = safe_unique_list(df_shapes["Shape"]) if "Shape" in df_shapes.columns else []
        col1, col2, col3 = st.columns(3)
        with col1:
            item = st.selectbox("Item", items) if items else ""
        with col2:
            zone = st.selectbox("House Zone", zones) if zones else ""
        with col3:
            sign_shape = st.session_state.get("selected_shape","")
            choices = [""] + shapes
            idx = choices.index(sign_shape) if sign_shape in choices else 0
            shape = st.selectbox("Shape (optional)", choices, index=idx)

        if item and zone:
            try:
                item_elem = df_elem_items.loc[df_elem_items["Item Name"]==item].iloc[0]["Element"]
            except Exception:
                item_elem = ""
            try:
                zone_elem = df_zones.loc[df_zones["Zone"]==zone].iloc[0]["Primary Element"]
            except Exception:
                zone_elem = ""
            zone_primary = str(zone_elem).split("+")[0].strip() if zone_elem else ""

            rel = "Neutral"
            try:
                row = df_rel[df_rel.iloc[:,0]==item_elem]
                if not row.empty and zone_primary in row.columns:
                    rel = row.iloc[0][zone_primary]
            except Exception:
                pass

            remedy = "OK / Supportive — keep as is"
            if zone == "Centre" and item_elem != "Space":
                remedy = "Keep centre open — relocate item to its best zones"
            elif isinstance(rel, str) and rel.startswith("Avoid"):
                if not df_pref.empty:
                    rowp = df_pref.loc[df_pref["Element"]==item_elem]
                    best = str(rowp.iloc[0]["Best Zones"]) if not rowp.empty else ""
                    remedy = f"Move to: {best}"
                else:
                    remedy = "Move to a better-suited zone for the item's element"
            elif rel == "Mild Avoid":
                remedy = "Balance with supportive colours/crystals of the zone element or relocate later"

            shape_elem = ""
            shape_rel = ""
            rec_shapes = ""
            if shape and not df_shapes.empty:
                srow = df_shapes.loc[df_shapes["Shape"]==shape]
                if not srow.empty:
                    shape_elem = str(srow.iloc[0]["Element"])
                    try:
                        rrow = df_rel[df_rel.iloc[:,0]==shape_elem]
                        if not rrow.empty and zone_primary in rrow.columns:
                            shape_rel = rrow.iloc[0][zone_primary]
                    except Exception:
                        shape_rel = ""
                rec_list = df_shapes.loc[df_shapes["Element"]==zone_primary]["Shape"].dropna().tolist() if zone_primary else []
                if rec_list:
                    rec_shapes = ", ".join(rec_list[:10])

            st.write(f"Item Element: {item_elem or '—'}   |   Zone Element (Primary): {zone_primary or '—'}")
            st.write(f"Verdict: {rel}")
            st.write(f"Remedy: {remedy}")
            if shape:
                st.write(f"Shape Element: {shape_elem or '—'}   |   Shape Verdict vs Zone: {shape_rel or '—'}")
            if rec_shapes:
                st.caption(f"Recommended shapes for this zone: {rec_shapes}")

# ===== Items Browser =====
with tabs[4]:
    st.subheader("Element Items Browser")
    if df_elem_items.empty:
        st.warning("Element_Items sheet missing.")
    else:
        elems = ["All"] + safe_unique_list(df_elem_items["Element"]) if "Element" in df_elem_items.columns else ["All"]
        cats  = ["All"] + safe_unique_list(df_elem_items["Category"]) if "Category" in df_elem_items.columns else ["All"]
        col1, col2 = st.columns(2)
        with col1:
            fe = st.selectbox("Element", elems, index=0)
        with col2:
            fc = st.selectbox("Category", cats, index=0)
        df_view = df_elem_items.copy()
        if "Element" in df_view.columns and fe != "All":
            df_view = df_view[df_view["Element"]==fe]
        if "Category" in df_view.columns and fc != "All":
            df_view = df_view[df_view["Category"]==fc]
        st.dataframe(df_view.reset_index(drop=True), use_container_width=True)
